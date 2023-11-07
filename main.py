from datetime import time
import time
import selenium
from selenium.webdriver.common.by import By
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
import openpyxl
from tkinter import ttk
import tkinter as tk

#Убирает пропуски и конвертирует всё в сумму
def library_converter(library):
    while True:
        try:
            new_library_summ=sum(library)
            return new_library_summ
        except TypeError:
            try:
                library.remove('')
            except ValueError:
                x_summ=0
                for i in library:
                    x_summ=x_summ+int(i)
                return x_summ

#Ищет пики значений, принимает любой список
def peak_finder(x_library):
    lier=0
    main_x=x_library[0]
    checker=True
    new_x=0
    new_x_library=[]
    for i in x_library:
        lier=lier+1
        if checker==True:
            print(x_library)
            new_x=int(new_x)+int(x_library[lier])
            if main_x-new_x==0:
                checker=False
                new_x_library.append(new_x)
                new_x=0
        else:
            checker=True
            try:
                main_x=int(x_library[lier])
            except IndexError:
                peak_summ=sum(new_x_library)
                return peak_summ
#Нахождение ключа в словаре по значению
def get_key(d, value):
    for k, v in d.items():
        if v == value:
            return k
#Сравнение дат(не используется, но лучше оставить)
def date_comparison(x_date,y_date):
    def dot_seeker(dot,ifer):
        if ifer == 1:
            dot_finder=dot.rfind('.')
            dot_findes=dot.find('.')
            dot = dot[dot_findes+1:dot_finder]
            try:
                dot=int(dot)
            except ValueError:
                dot=dot[1:]
        elif ifer == 0:
            dot_finder=dot.find('.')
            dot=dot[:dot_finder]
        try:
            dot = int(dot)
        except ValueError:
            dot = dot[1:]
        return dot
    day_x=dot_seeker(x_date,0)
    day_y=dot_seeker(y_date,0)
    month_x=dot_seeker(x_date,1)
    month_y=dot_seeker(y_date,1)
    if month_y>month_x:
        returner = True
        return returner
    elif month_y==month_x:
        if day_y>day_x:
            returner = True
            return returner
        elif day_y==day_x:
            returner = True
            return returner
        else:
            returner = False
            return returner
    else:
        returner = False
        return returner
#Поиск точки в строке
def dot_seeker(dot):
    dot_finder = dot.rfind('.')
    dot_findes = dot.find('.')
    dot = dot[dot_findes + 1:dot_finder]
    return dot
#Запись Excel файла
def write_file(create_number,create_date,status,full_price,product_name,send_to_house,get_on_house,purchase_price):
    work_table_name=comber.get()
    try:
        work_table = openpyxl.load_workbook(filename='Tables/'+work_table_name+'.xlsx')
        worksheet = work_table['Scaler_Place']
    except FileNotFoundError:
        work_table = openpyxl.Workbook()
        worksheet = work_table.create_sheet(index=0, title='Scaler_Place')
        worksheet['A1'] = 'Номер накладной'
        worksheet['B1'] = 'Дата создания'
        worksheet['C1'] = 'Статус'
        worksheet['D1'] = 'Общая стоимость'
        worksheet['E1'] = 'Наименование товара'
        worksheet['F1'] = 'Отправлено на склад'
        worksheet['G1'] = 'Принято на складе'
        worksheet['H1'] = 'Цена закупки'
    vals = []
    Inspector = 0
    x = ''
    while x != None:
        Inspector = Inspector + 1
        x = worksheet['A' + str(Inspector)].value
        vals.append(x)
    for Inspector in vals:
        if Inspector == create_number:
            worksheet['B' + str(vals.index(Inspector) + 1)] = create_date
            worksheet['C' + str(vals.index(Inspector) + 1)] = status
            worksheet['D' + str(vals.index(Inspector) + 1)] = full_price
            worksheet['E' + str(vals.index(Inspector) + 1)] = product_name
            worksheet['F' + str(vals.index(Inspector) + 1)] = send_to_house
            worksheet['G' + str(vals.index(Inspector) + 1)] = get_on_house
            worksheet['H' + str(vals.index(Inspector) + 1)] = purchase_price
            work_table.save('Tables/'+work_table_name+'.xlsx')
            return
        elif Inspector == None:
            worksheet['A' + str(vals.index(Inspector) + 1)] = create_number
            worksheet['B' + str(vals.index(Inspector) + 1)] = create_date
            worksheet['C' + str(vals.index(Inspector) + 1)] = status
            worksheet['D' + str(vals.index(Inspector) + 1)] = full_price
            worksheet['E' + str(vals.index(Inspector) + 1)] = product_name
            worksheet['F' + str(vals.index(Inspector) + 1)] = send_to_house
            worksheet['G' + str(vals.index(Inspector) + 1)] = get_on_house
            worksheet['H' + str(vals.index(Inspector) + 1)] = purchase_price
            work_table.save('Tables/'+work_table_name+'.xlsx')
def main(): #Главный скрипт по парсингу актов, запускается по кнопке
    #global checker_internet_var,comber,comber_date,checker_graphic_var (разобраться с работой init(сейчас прога не знает о существовании других данных))
    if checker_internet_var.get()==1: #Проверка на нажатие "Плохой интернет, замедляет программу в 2 раза
        multiplier=2
    elif checker_internet_var.get()==0:
        multiplier=1
    if checker_debugger_var.get()==1: #Проверка на нажатие"Debugging" уменьшает время ожидания от интерфейса до 15 секунд
        waiting=15
    elif checker_debugger_var.get()==0:
        waiting=900
    months_dict = {'12': 'Январь', '01': 'Февраль', '02': 'Март', '03': 'Апрель', '04': "Май", '05': "Июнь", "06": "Июль",
              "07": 'Август', '08': 'Сентябрь', '09': "Октябрь", '10': 'Ноябрь', '11': 'Декабрь'}
    i=0
    url=comber.get()
    if url=="TOPS":
        url='https://business.kazanexpress.ru/seller/4449/invoices/send'
    elif url=="Стельки":
        url='https://business.kazanexpress.ru/seller/65366/invoices/send'
    elif url=='Триколор':
        url='https://business.kazanexpress.ru/seller/10020/invoices/send'
    elif url=='Джибитсы':
        url='https://business.kazanexpress.ru/seller/51310/invoices/send'
    elif url=='Discont OFF':
        url='https://business.kazanexpress.ru/seller/10238/invoices/send'
    end_date = comber_date.get()
    end_month = get_key(months_dict,str(end_date) )
    try:
        service = Service(executable_path='chromedriver.exe')
        options = webdriver.ChromeOptions()
        if checker_graphic_var.get()==1: #Проверка на нажатие кнопки "Выключить графику"
            options.add_argument('--headless')
        driver = webdriver.Chrome(options=options)
    except:
        driver = webdriver.ChromiumEdge()
    driver.implicitly_wait(waiting)
    driver.get(url)
    field_finder=driver.find_element(By.XPATH, "//*[@id='username']")
    field_finder.send_keys('f4llno@yandex.ru')
    field_finder=driver.find_element(By.XPATH, "//*[@id='password']")
    field_finder.send_keys('Scaler_Password')
    time.sleep(1*multiplier)
    driver.find_element(By.XPATH, "//*[@id='signin']/section/div/section[2]/form/button").click()
    time.sleep(9*multiplier)
    driver.get(url)
    while True:
        element = driver.find_element(By.XPATH, "//*[@id='status-cell-" + str(i) + "-73-0']/div")
        driver.execute_script("arguments[0].scrollIntoView(true);", element)
        driver.find_element(By.XPATH, "//*[@id='status-cell-" + str(i) + "-73-0']/div").click()
        time.sleep(0.8)
        create_number=(driver.find_element(By.XPATH, "//*[@id='openInvoice']/div/div/div[2]/span[2]").text)
        print(create_number)
        create_date=driver.find_element(By.XPATH,'//*[@id="openInvoice"]/div/div/div[2]/span[4]').text
        status=driver.find_element(By.XPATH,'//*[@id="openInvoice"]/div/div/div[2]/span[6]').text
        if status=='Принята на складе':
            checker=True
            x=1
            temp_send_to_house=[]
            temp_get_on_house=[]
            while checker == True:
                driver.implicitly_wait(1)
                try:
                    try:
                        element=driver.find_element(By.XPATH,'//*[@id="openInvoice"]/div/div/div[3]/div[2]/div/table/tbody/tr[' + str(x) + ']/td[3]/div').text
                        driver.execute_script("arguments[0].scrollIntoView(true);", element)
                        temp_send_to_house.append(element)
                    except selenium.common.exceptions.JavascriptException:
                        temp_send_to_house.append(driver.find_element(By.XPATH,'//*[@id="openInvoice"]/div/div/div[3]/div[2]/div/table/tbody/tr[' + str(x) + ']/td[3]/div').text)
                        temp_get_on_house.append(driver.find_element(By.XPATH,'//*[@id="openInvoice"]/div/div/div[3]/div[2]/div/table/tbody/tr[' + str(x) + ']/td[4]/div').text)
                    x = int(x)
                    x = x + 1
                except selenium.common.exceptions.NoSuchElementException:
                    send_to_house=library_converter(temp_send_to_house)
                    get_on_house=library_converter(temp_get_on_house)
                    checker=False
            purchase_price=(driver.find_element(By.XPATH,'//*[@id="openInvoice"]/div/div/div[3]/div[2]/div/table/tbody/tr[1]/td[5]/div/span/em[1]').text)
        else:
            checker=True
            x=1
            temp_send_to_house=[]
            while checker == True:
                driver.implicitly_wait(1)
                try:
                    try:
                        element=driver.find_element(By.XPATH,'//*[@id="openInvoice"]/div/div/div[3]/div[2]/div/table/tbody/tr[' + str(x) + ']/td[3]/div').text
                        driver.execute_script("arguments[0].scrollIntoView(true);", element)
                        temp_send_to_house.append(element)
                    except selenium.common.exceptions.JavascriptException:
                        temp_send_to_house.append(driver.find_element(By.XPATH,'//*[@id="openInvoice"]/div/div/div[3]/div[2]/div/table/tbody/tr[' + str(x) + ']/td[3]/div').text)
                    x = int(x)
                    x = x + 1
                except selenium.common.exceptions.NoSuchElementException:
                    send_to_house=library_converter(temp_send_to_house)
                    checker=False
            purchase_price=(driver.find_element(By.XPATH,'//*[@id="openInvoice"]/div/div/div[3]/div[2]/div/table/tbody/tr[1]/td[4]/div/span/em[1]').text)
            get_on_house=0
        full_price=(driver.find_element(By.XPATH,'//*[@id="openInvoice"]/div/div/div[2]/span[8]').text)
        product_name=(driver.find_element(By.XPATH,'//*[@id="openInvoice"]/div/div/div[3]/div[2]/div/table/tbody/tr[1]/td[2]/div').text)
        create_month=dot_seeker(create_date)
        if create_month == end_month and status != 'Создана':
            return
        else:
            write_file(create_number,create_date,status,full_price,product_name,send_to_house,get_on_house,purchase_price)
            i=i+1
            driver.find_element(By.XPATH,'//*[@id="openInvoice"]/div/header/div[2]').click()
            time.sleep(0.3)

#Объявляются листы с магазинами и месяцами (используются в
magazines = ['TOPS', 'Стельки', 'Триколор', 'Джибитсы', 'Discont OFF']
months=['Январь','Февраль','Март','Апрель',"Май","Июнь","Июль",'Август','Сентябрь',"Октябрь",'Ноябрь','Декабрь']
# def init(): #Ниже инициализируется окно программы
window = tk.Tk()
window.title('Scaler')
window ['bg'] = 'gray10'
date_label = tk.Label(text="Выберите месяц начала парсинга",background='gray10',foreground='white')
date_label.pack()
comber_date = ttk.Combobox(values=months)
comber_date.pack()
magazine_label = tk.Label(text="Выберите магазин",background='gray10',foreground='white')
magazine_label.pack()
comber = ttk.Combobox(values=magazines)
comber.pack()
button = tk.Button(text='Начать выполнение', command=main)
button ['bg'] = 'white'
button.pack()
checker_internet_var=tk.IntVar()
checker_internet=ttk.Checkbutton(text='Плохой интернет',variable=checker_internet_var)
checker_internet.pack()
checker_graphic_var=tk.IntVar()
checker_graphic=ttk.Checkbutton(text='Отключить графику',variable=checker_graphic_var)
checker_graphic.pack()
checker_debugger_var=tk.IntVar()
checker_debugging=ttk.Checkbutton(text='Debugging',variable=checker_debugger_var)
checker_debugging.pack()
window.mainloop()
# if __name__=="__main__":
#     init()
