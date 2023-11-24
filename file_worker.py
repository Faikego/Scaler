import openpyxl
import os
#Запись Excel файла
def write_file(create_number,create_date,status,product_name,send_to_house,get_on_house,magazine):
    work_table_name=magazine
    try:
        work_table = openpyxl.load_workbook(filename='Tables/'+ work_table_name +'.xlsx')
        worksheet = work_table['Scaler_Place']
    except FileNotFoundError:
        try:
            os.mkdir("Tables")
            work_table = openpyxl.Workbook()
            worksheet = work_table.create_sheet(index=0, title='Scaler_Place')
            worksheet['A1'] = 'Номер накладной'
            worksheet['B1'] = 'Дата создания'
            worksheet['C1'] = 'Статус'
            worksheet['D1'] = 'Наименование товара'
            worksheet['E1'] = 'Отправлено на склад'
            worksheet['F1'] = 'Принято на складе'
        except FileExistsError:
            work_table = openpyxl.Workbook()
            worksheet = work_table.create_sheet(index=0, title='Scaler_Place')
            worksheet['A1'] = 'Номер накладной'
            worksheet['B1'] = 'Дата создания'
            worksheet['C1'] = 'Статус'
            worksheet['D1'] = 'Наименование товара'
            worksheet['E1'] = 'Отправлено на склад'
            worksheet['F1'] = 'Принято на складе'
    vals = []
    Inspector = 0
    x = ''
    while x != None:
        Inspector = Inspector + 1
        x = worksheet['A' + str(Inspector)].value
        vals.append(x)
    for Inspector in vals:
        if Inspector == create_number:
            worksheet['B' + str(vals.index(Inspector) + 1)] = create_date
            worksheet['C' + str(vals.index(Inspector) + 1)] = status
            worksheet['D' + str(vals.index(Inspector) + 1)] = product_name
            worksheet['E' + str(vals.index(Inspector) + 1)] = send_to_house
            worksheet['F' + str(vals.index(Inspector) + 1)] = get_on_house
            work_table.save('Tables/'+work_table_name+'.xlsx')
            return
        elif Inspector == None:
            worksheet['A' + str(vals.index(Inspector) + 1)] = create_number
            worksheet['B' + str(vals.index(Inspector) + 1)] = create_date
            worksheet['C' + str(vals.index(Inspector) + 1)] = status
            worksheet['D' + str(vals.index(Inspector) + 1)] = product_name
            worksheet['E' + str(vals.index(Inspector) + 1)] = send_to_house
            worksheet['F' + str(vals.index(Inspector) + 1)] = get_on_house
            work_table.save('Tables/'+work_table_name+'.xlsx')

#Убирает пропуски и конвертирует всё в сумму
def library_converter(library):
    #print(library)
    while True:
        try:
            library.remove('')
        except ValueError:
            #print(library)
            x_summ = 0
            for i in library:
                x_summ = x_summ + int(i)
            return x_summ
#Нахождение ключа в словаре по значению
def get_key(d, value):
    for k, v in d.items():
        if v == value:
            return k
#Поиск точки в строке
def dot_seeker(dot):
    dot_finder = dot.rfind('.')
    dot_findes = dot.find('.')
    #print(dot_findes,dot_finder)
    dot = dot[dot_findes + 1:dot_finder]
    #print(dot)
    return dot
def lopass_seeker(file):
    password_index=file.index(';')
    login=file[:password_index]
    password=file[password_index+1:]
    return login,password

def sorter_created(sort_library):
   output_library=[]
   counter=0
   for i in sort_library:
       if counter==1:
           output_library.append(i)
           counter=counter+1
       elif counter==2:
           counter=0
       else:
           counter=counter+1
   return output_library
def sorter_other(sort_library):
   get_on_house_library=[]
   send_to_house_library=[]
   counter=0
   for i in sort_library:
       if counter==1:
           send_to_house_library.append(i)
           counter=counter+1
       elif counter==2:
           get_on_house_library.append(i)
           counter = counter + 1
       elif counter==3:
           counter=0
       else:
           counter=counter+1
   return send_to_house_library,get_on_house_library
def zero_library_creater(library):
    output_library=[]
    for i in library:
        output_library.append(0)
    return output_library